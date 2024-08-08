from flask import Flask, jsonify, request, render_template, redirect, url_for
from openpyxl import Workbook, load_workbook
import os
import json

app = Flask(__name__)

REVIEW_FILE = 'post.xlsx'
LOGIN_CREDENTIALS_FILE = 'credentials.xlsx'

# Initialize Excel files if they don't exist
if not os.path.exists(LOGIN_CREDENTIALS_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Emailid', 'Password', 'Name', 'Role'])
    workbook.save(LOGIN_CREDENTIALS_FILE)

if not os.path.exists(REVIEW_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Restaurant Name', 'User Name', 'Review', 'Recommended Dishes', 'Suitable For', 'Safety Rating', 'Food Rating', 'Overall Rating'])
    workbook.save(REVIEW_FILE)

@app.route('/')
def intro():
    return render_template('intro.html')

@app.route('/home')
def home():
    return render_template('home.html')

@app.route('/post')
def post():
    return render_template('post.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user_mail = request.form['Email']
        user_password = request.form['Password']
        workbook = load_workbook(LOGIN_CREDENTIALS_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == user_mail and row[1] == user_password:
                return redirect(url_for('home'))
        return render_template('login.html', error="Invalid credentials")
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']

        # Save the new user to the Excel file
        workbook = load_workbook(LOGIN_CREDENTIALS_FILE)
        sheet = workbook.active
        sheet.append([email, password, name, role])
        workbook.save(LOGIN_CREDENTIALS_FILE)

        return redirect(url_for('login'))
    
    return render_template('signup.html')

@app.route('/launch')
def launch():
    return render_template('launch.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/blog')
def blog():
    return render_template('blog.html')

@app.route('/offers')
def offers():
    return render_template('offers.html')

@app.route('/singles')
def singles():
    return render_template('singles.html')

@app.route('/couples')
def couples():
    return render_template('couples.html')

@app.route('/midnight')
def midnight():
    return render_template('midnight.html')

@app.route('/pubs')
def pubs():
    return render_template('pubs.html')

@app.route('/postr', methods=['GET','POST'])
def postr():
    if request.method == 'POST':
        hotel_name = request.form['hotelName']
        user_name = request.form['userName']
        review = request.form['review']
        recommended_dishes = request.form['recommendedDishes']
        suitability = request.form['suitability']
        safety_rating = request.form['safetyRating']
        food_rating = request.form['foodRating']
        overall_rating = request.form['overallRating']

        # Load the Excel file
        workbook = load_workbook(REVIEW_FILE)
        sheet = workbook.active

        # Append the new review to the Excel file
        sheet.append([hotel_name, user_name, review, recommended_dishes, suitability, safety_rating, food_rating, overall_rating])
        workbook.save(REVIEW_FILE)

        return redirect(url_for('post'))  # Redirect to the same page after posting

    return render_template('post.html')

@app.route('/view_reviews')
def view_reviews():
    workbook = load_workbook(REVIEW_FILE)
    sheet = workbook.active
    reviews = []
    
    # Skip the header row and read the data
    for row in sheet.iter_rows(min_row=2, values_only=True):
        reviews.append({
            'restaurant_name': row[0],
            'user_name': row[1],
            'review': row[2],
            'recommended_dishes': row[3],
            'suitability': row[4],
            'safety_rating': row[5],
            'food_rating': row[6],
            'overall_rating': row[7]
        })
    
    return render_template('view.html', reviews=reviews)

@app.route('/api/reviews', methods=['GET'])
def get_reviews():
    if os.path.exists(REVIEW_FILE):
        workbook = load_workbook(REVIEW_FILE)
        sheet = workbook.active
        reviews = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            review = {
                'hotelName': row[0],
                'userName': row[1],
                'review': row[2],
                'recommendedDishes': row[3],
                'suitability': row[4],
                'safetyRating': row[5],
                'foodRating': row[6],
                'overallRating': row[7]
            }
            reviews.append(review)
        return jsonify(reviews)
    return jsonify([])


if __name__ == "__main__":
    app.run(debug=True)
